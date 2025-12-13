"""
Report generation and formatting for creator analysis
"""

import json
from typing import Dict, Optional
from datetime import datetime
import pandas as pd
import io

from storage import get_db

# Import export libraries at module level
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
    print("[DEBUG] reportlab imported successfully")
except ImportError as e:
    REPORTLAB_AVAILABLE = False
    print(f"[WARN] reportlab import failed: {e}")
    import sys
    print(f"[DEBUG] sys.path: {sys.path}")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
    print("[DEBUG] openpyxl imported successfully")
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    print(f"[WARN] openpyxl import failed: {e}")

try:
    import markdown
    MARKDOWN_AVAILABLE = True
except ImportError:
    MARKDOWN_AVAILABLE = False


class ReportGenerator:
    """
    Creates formatted reports with:
    - Platform statistics table
    - Demographics breakdown
    - Content analysis
    - Brand fit scoring
    - Recommendations
    """

    def __init__(self):
        self.db = get_db()

    def generate_report(
        self,
        creator_id: int,
        brief_id: int,
        format: str = "markdown"
    ) -> str:
        """
        Generate comprehensive creator analysis report

        Args:
            creator_id: Creator ID
            brief_id: Brief ID
            format: Output format (markdown, html, text)

        Returns:
            Formatted report string
        """
        # Get creator and brief info
        creator = self.db.get_creator(creator_id)
        brief = self.db.get_brief(brief_id)
        report = self.db.get_creator_report(brief_id, creator_id)

        if not creator or not brief or not report:
            return "Report not found"

        # Get social accounts and analytics
        accounts_df = self.db.get_social_accounts(creator_id)

        platform_data = []
        demographics_data = {}
        posts_data = []
        for _, account in accounts_df.iterrows():
            analytics = self.db.get_latest_analytics(account['id'])
            if analytics:
                platform_data.append({
                    'platform': account['platform'].title(),
                    'handle': account['handle'] or 'N/A',
                    'followers': analytics.get('followers_count', 0),
                    'total_posts': analytics.get('total_posts', 0),
                    'engagement_rate': analytics.get('engagement_rate', 0.0),
                    'profile_url': account['profile_url']
                })

            # Get demographics data for this platform
            demographics = self.db.get_demographics_data(account['id'])
            if demographics:
                demographics_data[account['platform']] = demographics

            # Get posts analyzed for this platform
            posts_df = self.db.get_posts_for_account(account['id'], limit=50)
            if not posts_df.empty:
                for _, post in posts_df.iterrows():
                    posts_data.append({
                        'platform': account['platform'].title(),
                        'post_date': post.get('post_date'),
                        'post_type': post.get('post_type', 'post'),
                        'caption': post.get('caption', ''),
                        'likes': post.get('likes_count', 0),
                        'comments': post.get('comments_count', 0),
                        'shares': post.get('shares_count', 0),
                        'views': post.get('views_count', 0),
                        'duration': post.get('duration_seconds', 0),
                        'post_url': post.get('post_url', ''),
                        'content_themes': post.get('content_themes', ''),
                        'brand_safety_score': post.get('brand_safety_score', 0.0),
                        'sentiment_score': post.get('sentiment_score', 0.0)
                    })

        if format == "markdown":
            return self._generate_markdown_report(
                creator, brief, report, platform_data, demographics_data, posts_data
            )
        elif format == "html":
            return self._generate_html_report(
                creator, brief, report, platform_data, demographics_data, posts_data
            )
        else:
            return self._generate_text_report(
                creator, brief, report, platform_data, demographics_data, posts_data
            )

    def _generate_markdown_report(
        self,
        creator: Dict,
        brief: Dict,
        report: Dict,
        platform_data: list,
        demographics_data: Dict = None,
        posts_data: list = None
    ) -> str:
        """Generate markdown formatted report"""

        # Platform statistics table
        platform_table = "| Platform | Handle | Followers | Total Posts | Profile |\n"
        platform_table += "|----------|--------|-----------|-------------|----------|\n"

        for p in platform_data:
            platform_table += f"| {p['platform']} | {p['handle']} | {p['followers']:,} | {p['total_posts']:,} | [Link]({p['profile_url']}) |\n"

        # Demographics section
        demographics_section = ""
        if demographics_data:
            demographics_section = "\n## Audience Demographics\n\n"
            for platform, demo in demographics_data.items():
                demographics_section += f"### {platform.title()}\n\n"

                # Gender distribution
                if 'gender' in demo and demo['gender']:
                    demographics_section += "**Gender Distribution:**\n"
                    for gender, percentage in demo['gender'].items():
                        demographics_section += f"- {gender.title()}: {percentage}%\n"
                    demographics_section += "\n"

                # Age brackets
                if 'age_brackets' in demo and demo['age_brackets']:
                    demographics_section += "**Age Distribution:**\n"
                    for age_range, percentage in demo['age_brackets'].items():
                        demographics_section += f"- {age_range}: {percentage}%\n"
                    demographics_section += "\n"

                # Geography
                if 'geography' in demo and demo['geography']:
                    demographics_section += "**Geographic Distribution:**\n"
                    for geo in demo['geography'][:5]:  # Top 5 countries
                        demographics_section += f"- {geo.get('country', 'Unknown')}: {geo.get('percentage', 0)}%\n"
                    demographics_section += "\n"

                # Languages
                if 'languages' in demo and demo['languages']:
                    demographics_section += "**Languages:**\n"
                    for lang in demo['languages'][:3]:  # Top 3 languages
                        demographics_section += f"- {lang.get('language', 'Unknown')}: {lang.get('percentage', 0)}%\n"
                    demographics_section += "\n"

                # Data confidence
                if 'data_confidence' in demo:
                    demographics_section += f"*Data Confidence: {demo['data_confidence'].title()}*\n"
                if 'data_source' in demo:
                    demographics_section += f"*Source: {demo['data_source'].replace('_', ' ').title()}*\n"
                demographics_section += "\n---\n\n"

        # Posts analysis section
        posts_section = ""
        if posts_data:
            posts_section = "\n## Recent Posts Analyzed\n\n"
            posts_section += f"**Total Posts Analyzed**: {len(posts_data)}\n\n"

            # Group posts by platform
            posts_by_platform = {}
            for post in posts_data:
                platform = post['platform']
                if platform not in posts_by_platform:
                    posts_by_platform[platform] = []
                posts_by_platform[platform].append(post)

            # Display posts per platform
            for platform, platform_posts in posts_by_platform.items():
                posts_section += f"### {platform} ({len(platform_posts)} posts)\n\n"

                # Calculate aggregate metrics
                total_likes = sum(p.get('likes', 0) for p in platform_posts)
                total_comments = sum(p.get('comments', 0) for p in platform_posts)
                total_views = sum(p.get('views', 0) for p in platform_posts if p.get('views'))
                avg_brand_safety = sum(p.get('brand_safety_score', 0) for p in platform_posts if p.get('brand_safety_score')) / len([p for p in platform_posts if p.get('brand_safety_score')]) if any(p.get('brand_safety_score') for p in platform_posts) else 0

                posts_section += f"**Aggregate Metrics:**\n"
                posts_section += f"- Total Likes: {total_likes:,}\n"
                posts_section += f"- Total Comments: {total_comments:,}\n"
                if total_views > 0:
                    posts_section += f"- Total Views: {total_views:,}\n"
                if avg_brand_safety > 0:
                    posts_section += f"- Average Brand Safety Score: {avg_brand_safety:.1f}/5.0\n"
                posts_section += "\n"

                # Show top 5 posts by engagement
                sorted_posts = sorted(platform_posts, key=lambda p: p.get('likes', 0) + p.get('comments', 0) * 3, reverse=True)[:5]
                posts_section += "**Top 5 Posts by Engagement:**\n\n"

                for idx, post in enumerate(sorted_posts, 1):
                    posts_section += f"**{idx}.** "

                    # Post date
                    if post.get('post_date'):
                        try:
                            post_date = post['post_date'].split('T')[0] if isinstance(post['post_date'], str) else str(post['post_date'])[:10]
                            posts_section += f"*{post_date}* - "
                        except:
                            pass

                    # Post type
                    if post.get('post_type'):
                        posts_section += f"`{post['post_type']}` "

                    # Caption preview
                    caption = post.get('caption', '')
                    if caption:
                        caption_preview = caption[:100] + "..." if len(caption) > 100 else caption
                        posts_section += f"{caption_preview}\n"
                    else:
                        posts_section += "\n"

                    # Engagement metrics
                    metrics = []
                    if post.get('likes'):
                        metrics.append(f"{post['likes']:,} likes")
                    if post.get('comments'):
                        metrics.append(f"{post['comments']:,} comments")
                    if post.get('views'):
                        metrics.append(f"{post['views']:,} views")
                    if post.get('shares'):
                        metrics.append(f"{post['shares']:,} shares")

                    if metrics:
                        posts_section += f"   - {' | '.join(metrics)}\n"

                    # Brand safety score
                    if post.get('brand_safety_score'):
                        posts_section += f"   - Brand Safety: {post['brand_safety_score']:.1f}/5.0\n"

                    # Content themes
                    if post.get('content_themes'):
                        themes = post['content_themes']
                        if isinstance(themes, str):
                            try:
                                import json
                                themes = json.loads(themes)
                            except:
                                themes = [themes]
                        if themes and (isinstance(themes, list) and themes):
                            themes_str = ', '.join(themes[:3]) if isinstance(themes, list) else str(themes)
                            posts_section += f"   - Themes: {themes_str}\n"

                    # Post URL
                    if post.get('post_url'):
                        posts_section += f"   - [View Post]({post['post_url']})\n"

                    posts_section += "\n"

                posts_section += "---\n\n"

        # Content themes section
        content_analysis = report.get('content_analysis', {})
        content_themes = content_analysis.get('content_themes', [])
        content_themes_section = ""
        if content_themes:
            content_themes_section = "\n## Content Themes\n\n"
            content_themes_section += "Primary topics and themes identified across analyzed content:\n\n"
            for theme in content_themes[:8]:  # Show top 8 themes
                content_themes_section += f"- {theme}\n"
            content_themes_section += "\n"

            # Add key metrics from content analysis
            if content_analysis:
                content_themes_section += "**Content Quality Metrics:**\n\n"
                if content_analysis.get('brand_safety_score'):
                    content_themes_section += f"- Brand Safety Score: {content_analysis.get('brand_safety_score', 'N/A')}/5.0\n"
                if content_analysis.get('authenticity_score'):
                    content_themes_section += f"- Authenticity Score: {content_analysis.get('authenticity_score', 'N/A')}/5.0\n"
                if content_analysis.get('audience_engagement_quality'):
                    content_themes_section += f"- Audience Engagement: {content_analysis.get('audience_engagement_quality', 'N/A').title()}\n"
                if content_analysis.get('production_quality'):
                    content_themes_section += f"- Production Quality: {content_analysis.get('production_quality', 'N/A').title()}\n"
                if content_analysis.get('sentiment'):
                    content_themes_section += f"- Overall Sentiment: {content_analysis.get('sentiment', 'N/A').title()}\n"
                content_themes_section += "\n---\n\n"

        # Strengths and concerns
        strengths = report.get('strengths', [])
        concerns = report.get('concerns', [])
        recommendations = report.get('recommendations', [])

        strengths_list = '\n'.join([f"- {s}" for s in strengths]) if strengths else "- None identified"
        concerns_list = '\n'.join([f"- {c}" for c in concerns]) if concerns else "- None identified"
        recommendations_list = '\n'.join([f"- {r}" for r in recommendations]) if recommendations else "- None provided"

        # Video insights section
        video_insights = report.get('video_insights', [])
        video_section = ""
        if video_insights:
            video_section = "\n## Video Analysis\n\n"
            video_section += f"**{len(video_insights)} videos analyzed**\n\n"
            for idx, video in enumerate(video_insights, 1):
                video_section += f"### Video {idx}: {video.get('title', 'Unknown')}\n\n"
                video_section += f"- **URL**: [{video.get('url', 'N/A')}]({video.get('url', '#')})\n"
                video_section += f"- **Analysis Method**: {video.get('analysis_method', 'N/A').title()}\n"
                video_section += f"- **Brand Safety Score**: {video.get('brand_safety_score', 'N/A')}/5.0\n"
                if video.get('relevance_score'):
                    video_section += f"- **Content Relevance**: {video.get('relevance_score', 'N/A')}/5.0\n"
                if video.get('key_topics'):
                    video_section += f"- **Key Topics**: {', '.join(video.get('key_topics', []))}\n"
                if video.get('concerns'):
                    concerns_str = ', '.join(video.get('concerns', []))
                    video_section += f"- **Concerns**: {concerns_str}\n"
                video_section += "\n"
            video_section += "---\n\n"

        # Generate report
        markdown = f"""# Creator Analysis Report: {creator['name']}

**Brief**: {brief['name']}
**Date**: {datetime.now().strftime('%B %d, %Y')}
**Overall Brand Fit**: {report['overall_score']}/5.0

---

## Executive Summary

{report['summary']}

---

## Platform Statistics

{platform_table}

---

{demographics_section}{posts_section}{content_themes_section}## Content Analysis

### Strengths
{strengths_list}

### Concerns
{concerns_list}

---

{video_section}## Recommendations

{recommendations_list}

---

## Brand Context

{brief.get('brand_context', 'No brand context provided')}

---

*Generated with BrandSafe Talent Analysis Tool*
*Report ID: {report['id']} | Model: {report.get('model_used', 'N/A')} | Cost: ${report.get('analysis_cost', 0.0):.4f}*
"""

        return markdown

    def _generate_html_report(
        self,
        creator: Dict,
        brief: Dict,
        report: Dict,
        platform_data: list,
        demographics_data: Dict = None,
        posts_data: list = None
    ) -> str:
        """Generate HTML formatted report"""

        # Platform statistics table
        platform_rows = ""
        for p in platform_data:
            platform_rows += f"""
                <tr>
                    <td>{p['platform']}</td>
                    <td>{p['handle']}</td>
                    <td>{p['followers']:,}</td>
                    <td>{p['total_posts']:,}</td>
                    <td><a href="{p['profile_url']}" target="_blank">View Profile</a></td>
                </tr>
            """

        # Posts analysis section HTML
        posts_html = ""
        if posts_data:
            posts_html = f'<h2>Recent Posts Analyzed</h2><p><strong>Total Posts Analyzed:</strong> {len(posts_data)}</p>'

            # Group posts by platform
            posts_by_platform = {}
            for post in posts_data:
                platform = post['platform']
                if platform not in posts_by_platform:
                    posts_by_platform[platform] = []
                posts_by_platform[platform].append(post)

            for platform, platform_posts in posts_by_platform.items():
                posts_html += f'<h3>{platform} ({len(platform_posts)} posts)</h3>'

                # Aggregate metrics
                total_likes = sum(p.get('likes', 0) for p in platform_posts)
                total_comments = sum(p.get('comments', 0) for p in platform_posts)
                total_views = sum(p.get('views', 0) for p in platform_posts if p.get('views'))

                posts_html += '<p><strong>Aggregate Metrics:</strong></p><ul>'
                posts_html += f'<li>Total Likes: {total_likes:,}</li>'
                posts_html += f'<li>Total Comments: {total_comments:,}</li>'
                if total_views > 0:
                    posts_html += f'<li>Total Views: {total_views:,}</li>'
                posts_html += '</ul>'

                # Top 5 posts
                sorted_posts = sorted(platform_posts, key=lambda p: p.get('likes', 0) + p.get('comments', 0) * 3, reverse=True)[:5]
                posts_html += '<p><strong>Top 5 Posts by Engagement:</strong></p>'

                for idx, post in enumerate(sorted_posts, 1):
                    posts_html += f'<div style="margin: 15px 0; padding: 10px; background-color: #f8f9fa; border-radius: 5px;">'
                    posts_html += f'<p><strong>{idx}.</strong> '

                    # Post date and type
                    if post.get('post_date'):
                        try:
                            post_date = post['post_date'].split('T')[0] if isinstance(post['post_date'], str) else str(post['post_date'])[:10]
                            posts_html += f'<em>{post_date}</em> - '
                        except:
                            pass

                    if post.get('post_type'):
                        posts_html += f'<code>{post["post_type"]}</code> '

                    posts_html += '</p>'

                    # Caption
                    caption = post.get('caption', '')
                    if caption:
                        caption_preview = caption[:150] + "..." if len(caption) > 150 else caption
                        posts_html += f'<p>{caption_preview}</p>'

                    # Metrics
                    posts_html += '<ul style="margin: 5px 0;">'
                    if post.get('likes'):
                        posts_html += f'<li>{post["likes"]:,} likes</li>'
                    if post.get('comments'):
                        posts_html += f'<li>{post["comments"]:,} comments</li>'
                    if post.get('views'):
                        posts_html += f'<li>{post["views"]:,} views</li>'
                    if post.get('brand_safety_score'):
                        posts_html += f'<li>Brand Safety: {post["brand_safety_score"]:.1f}/5.0</li>'
                    if post.get('post_url'):
                        posts_html += f'<li><a href="{post["post_url"]}" target="_blank">View Post</a></li>'
                    posts_html += '</ul></div>'

        # Strengths, concerns, recommendations
        strengths = report.get('strengths', [])
        concerns = report.get('concerns', [])
        recommendations = report.get('recommendations', [])

        strengths_html = '<ul>' + ''.join([f'<li>{s}</li>' for s in strengths]) + '</ul>' if strengths else '<p>None identified</p>'
        concerns_html = '<ul>' + ''.join([f'<li>{c}</li>' for c in concerns]) + '</ul>' if concerns else '<p>None identified</p>'
        recommendations_html = '<ul>' + ''.join([f'<li>{r}</li>' for r in recommendations]) + '</ul>' if recommendations else '<p>None provided</p>'

        html = f"""
<!DOCTYPE html>
<html>
<head>
    <title>Creator Analysis Report: {creator['name']}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            line-height: 1.6;
        }}
        h1 {{ color: #333; }}
        h2 {{ color: #555; border-bottom: 2px solid #eee; padding-bottom: 10px; }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            background-color: #f8f9fa;
            font-weight: bold;
        }}
        .score {{
            font-size: 24px;
            font-weight: bold;
            color: #28a745;
        }}
        .metadata {{
            color: #666;
            font-size: 14px;
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #eee;
        }}
        .summary {{
            background-color: #f8f9fa;
            padding: 20px;
            border-radius: 5px;
            margin: 20px 0;
        }}
    </style>
</head>
<body>
    <h1>Creator Analysis Report: {creator['name']}</h1>

    <p><strong>Brief:</strong> {brief['name']}<br>
    <strong>Date:</strong> {datetime.now().strftime('%B %d, %Y')}<br>
    <strong>Overall Brand Fit:</strong> <span class="score">{report['overall_score']}/10</span></p>

    <hr>

    <h2>Executive Summary</h2>
    <div class="summary">
        {report['summary'].replace('\n', '<br>')}
    </div>

    <h2>Platform Statistics</h2>
    <table>
        <thead>
            <tr>
                <th>Platform</th>
                <th>Handle</th>
                <th>Followers</th>
                <th>Total Posts</th>
                <th>Profile</th>
            </tr>
        </thead>
        <tbody>
            {platform_rows}
        </tbody>
    </table>

    {posts_html}

    <h2>Content Analysis</h2>

    <h3>Strengths</h3>
    {strengths_html}

    <h3>Concerns</h3>
    {concerns_html}

    <h2>Recommendations</h2>
    {recommendations_html}

    <h2>Brand Context</h2>
    <p>{brief.get('brand_context', 'No brand context provided')}</p>

    <div class="metadata">
        <p><em>Generated with BrandSafe Talent Analysis Tool</em><br>
        Report ID: {report['id']} | Model: {report.get('model_used', 'N/A')} | Cost: ${report.get('analysis_cost', 0.0):.2f}</p>
    </div>
</body>
</html>
"""

        return html

    def _generate_text_report(
        self,
        creator: Dict,
        brief: Dict,
        report: Dict,
        platform_data: list,
        demographics_data: Dict = None,
        posts_data: list = None
    ) -> str:
        """Generate plain text formatted report"""

        text = f"""
{'='*80}
CREATOR ANALYSIS REPORT: {creator['name']}
{'='*80}

Brief: {brief['name']}
Date: {datetime.now().strftime('%B %d, %Y')}
Overall Brand Fit: {report['overall_score']}/10

{'-'*80}
EXECUTIVE SUMMARY
{'-'*80}

{report['summary']}

{'-'*80}
PLATFORM STATISTICS
{'-'*80}

"""

        for p in platform_data:
            text += f"""
Platform: {p['platform']}
Handle: {p['handle']}
Followers: {p['followers']:,}
Total Posts: {p['total_posts']:,}
Profile: {p['profile_url']}
"""

        # Posts analysis section
        if posts_data:
            text += f"""
{'-'*80}
RECENT POSTS ANALYZED
{'-'*80}

Total Posts Analyzed: {len(posts_data)}

"""
            # Group posts by platform
            posts_by_platform = {}
            for post in posts_data:
                platform = post['platform']
                if platform not in posts_by_platform:
                    posts_by_platform[platform] = []
                posts_by_platform[platform].append(post)

            for platform, platform_posts in posts_by_platform.items():
                text += f"\n{platform} ({len(platform_posts)} posts)\n"
                text += "-" * 40 + "\n\n"

                # Aggregate metrics
                total_likes = sum(p.get('likes', 0) for p in platform_posts)
                total_comments = sum(p.get('comments', 0) for p in platform_posts)
                total_views = sum(p.get('views', 0) for p in platform_posts if p.get('views'))

                text += "Aggregate Metrics:\n"
                text += f"  - Total Likes: {total_likes:,}\n"
                text += f"  - Total Comments: {total_comments:,}\n"
                if total_views > 0:
                    text += f"  - Total Views: {total_views:,}\n"
                text += "\n"

                # Top 5 posts
                sorted_posts = sorted(platform_posts, key=lambda p: p.get('likes', 0) + p.get('comments', 0) * 3, reverse=True)[:5]
                text += "Top 5 Posts by Engagement:\n\n"

                for idx, post in enumerate(sorted_posts, 1):
                    text += f"{idx}. "

                    # Post date
                    if post.get('post_date'):
                        try:
                            post_date = post['post_date'].split('T')[0] if isinstance(post['post_date'], str) else str(post['post_date'])[:10]
                            text += f"{post_date} - "
                        except:
                            pass

                    # Post type
                    if post.get('post_type'):
                        text += f"[{post['post_type']}] "

                    # Caption
                    caption = post.get('caption', '')
                    if caption:
                        caption_preview = caption[:80] + "..." if len(caption) > 80 else caption
                        text += f"\n   {caption_preview}\n"
                    else:
                        text += "\n"

                    # Metrics
                    if post.get('likes'):
                        text += f"   - Likes: {post['likes']:,}\n"
                    if post.get('comments'):
                        text += f"   - Comments: {post['comments']:,}\n"
                    if post.get('views'):
                        text += f"   - Views: {post['views']:,}\n"
                    if post.get('brand_safety_score'):
                        text += f"   - Brand Safety: {post['brand_safety_score']:.1f}/5.0\n"
                    if post.get('post_url'):
                        text += f"   - URL: {post['post_url']}\n"

                    text += "\n"

        text += f"""
{'-'*80}
CONTENT ANALYSIS
{'-'*80}

Strengths:
"""
        for s in report.get('strengths', []):
            text += f"  - {s}\n"

        text += "\nConcerns:\n"
        for c in report.get('concerns', []):
            text += f"  - {c}\n"

        text += f"""
{'-'*80}
RECOMMENDATIONS
{'-'*80}

"""
        for r in report.get('recommendations', []):
            text += f"  - {r}\n"

        text += f"""
{'-'*80}
BRAND CONTEXT
{'-'*80}

{brief.get('brand_context', 'No brand context provided')}

{'-'*80}

Generated with BrandSafe Talent Analysis Tool
Report ID: {report['id']} | Model: {report.get('model_used', 'N/A')} | Cost: ${report.get('analysis_cost', 0.0):.2f}

{'='*80}
"""

        return text

    def export_report_to_file(
        self,
        creator_id: int,
        brief_id: int,
        format: str = "markdown",
        filename: Optional[str] = None
    ) -> str:
        """
        Export report to file

        Args:
            creator_id: Creator ID
            brief_id: Brief ID
            format: Output format
            filename: Optional custom filename

        Returns:
            Path to exported file
        """
        report_content = self.generate_report(creator_id, brief_id, format)

        if not filename:
            creator = self.db.get_creator(creator_id)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = "md" if format == "markdown" else "html" if format == "html" else "txt"
            filename = f"report_{creator['name'].replace(' ', '_')}_{timestamp}.{ext}"

        filepath = f"./data/exports/{filename}"

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(report_content)

        return filepath

    def generate_brief_report(
        self,
        brief_id: int,
        format: str = "markdown"
    ) -> str:
        """
        Generate comprehensive report for all creators in a brief

        Args:
            brief_id: Brief ID
            format: Output format (markdown only for now)

        Returns:
            Formatted complete brief report
        """
        # Get brief info
        brief = self.db.get_brief(brief_id)
        if not brief:
            return "Brief not found"

        # Get all reports for this brief
        reports_df = self.db.get_reports_for_brief(brief_id)

        if reports_df.empty:
            return "No reports found for this brief"

        # Sort by score descending
        reports_df = reports_df.sort_values('overall_score', ascending=False)

        # Calculate aggregate statistics
        avg_score = reports_df['overall_score'].mean()
        total_cost = reports_df['analysis_cost'].sum()
        top_creators = reports_df.head(3)

        # Start building the report
        markdown = f"""# {brief['name']} - Complete Analysis Report

**Generated:** {datetime.now().strftime("%B %d, %Y at %I:%M %p")}

---

## Executive Summary

**Brief Description:** {brief.get('description', 'N/A')}

**Brand Context:** {brief.get('brand_context', 'N/A')}

### Key Metrics
- **Total Creators Analyzed:** {len(reports_df)}
- **Average Brand Fit Score:** {avg_score:.1f}/5.0
- **Total Analysis Cost:** ${total_cost:.2f}

### Top 3 Recommended Creators
"""

        for idx, (_, row) in enumerate(top_creators.iterrows(), 1):
            score_emoji = "🟢" if row['overall_score'] >= 4.0 else "🟡" if row['overall_score'] >= 3.0 else "🔴"
            markdown += f"{idx}. {score_emoji} **{row['creator_name']}** - {row['overall_score']:.1f}/5.0 ({row['primary_platform']})\n"

        markdown += "\n---\n\n## Creator Comparison Table\n\n"
        markdown += "| Rank | Creator | Platform | Brand Fit | Followers | Cost |\n"
        markdown += "|------|---------|----------|-----------|-----------|------|\n"

        for idx, (_, row) in enumerate(reports_df.iterrows(), 1):
            # Get platform stats for follower count
            creator = self.db.get_creator(int(row['creator_id']))
            accounts_df = self.db.get_social_accounts(int(row['creator_id']))

            total_followers = 0
            if not accounts_df.empty:
                for _, acc in accounts_df.iterrows():
                    analytics = self.db.get_latest_analytics(int(acc['id']))
                    if analytics:
                        total_followers += analytics.get('followers_count', 0)

            followers_display = f"{total_followers:,}" if total_followers > 0 else "N/A"

            markdown += f"| {idx} | {row['creator_name']} | {row['primary_platform']} | {row['overall_score']:.1f}/5.0 | {followers_display} | ${row['analysis_cost']:.2f} |\n"

        markdown += "\n---\n\n## Individual Creator Reports\n\n"

        # Add each creator's detailed report
        for idx, (_, row) in enumerate(reports_df.iterrows(), 1):
            score_emoji = "🟢" if row['overall_score'] >= 4.0 else "🟡" if row['overall_score'] >= 3.0 else "🔴"

            markdown += f"### {idx}. {score_emoji} {row['creator_name']} - {row['overall_score']:.1f}/5.0\n\n"

            # Parse stored report data
            try:
                strengths = json.loads(row['strengths']) if isinstance(row['strengths'], str) else row['strengths']
                concerns = json.loads(row['concerns']) if isinstance(row['concerns'], str) else row['concerns']
                recommendations = json.loads(row['recommendations']) if isinstance(row['recommendations'], str) else row['recommendations']
            except:
                strengths = []
                concerns = []
                recommendations = []

            # Summary
            markdown += f"**Summary:** {row.get('summary', 'No summary available')}\n\n"

            # Strengths
            if strengths:
                markdown += "**Strengths:**\n"
                for strength in strengths:
                    markdown += f"- {strength}\n"
                markdown += "\n"

            # Concerns
            if concerns:
                markdown += "**Potential Concerns:**\n"
                for concern in concerns:
                    markdown += f"- {concern}\n"
                markdown += "\n"

            # Recommendations
            if recommendations:
                markdown += "**Recommendations:**\n"
                for rec in recommendations:
                    markdown += f"- {rec}\n"
                markdown += "\n"

            # Get content analysis for themes
            try:
                content_analysis = json.loads(row['content_analysis']) if isinstance(row['content_analysis'], str) else row.get('content_analysis', {})
                if content_analysis and content_analysis.get('content_themes'):
                    content_themes = content_analysis.get('content_themes', [])
                    markdown += "**Content Themes:**\n"
                    for theme in content_themes[:8]:
                        markdown += f"- {theme}\n"
                    markdown += "\n"

                    # Add quality metrics
                    markdown += "**Content Quality Metrics:**\n"
                    if content_analysis.get('brand_safety_score'):
                        markdown += f"- Brand Safety: {content_analysis.get('brand_safety_score')}/5.0\n"
                    if content_analysis.get('authenticity_score'):
                        markdown += f"- Authenticity: {content_analysis.get('authenticity_score')}/5.0\n"
                    if content_analysis.get('audience_engagement_quality'):
                        markdown += f"- Engagement Quality: {content_analysis.get('audience_engagement_quality', 'N/A').title()}\n"
                    markdown += "\n"
            except:
                pass

            # Platform details with demographics and posts
            creator_id = int(row['creator_id'])
            accounts_df = self.db.get_social_accounts(creator_id)

            if not accounts_df.empty:
                markdown += "**Platform Presence:**\n\n"

                for _, acc in accounts_df.iterrows():
                    platform = acc['platform']
                    markdown += f"#### {platform.title()}\n\n"

                    analytics = self.db.get_latest_analytics(int(acc['id']))
                    if analytics:
                        followers = f"{analytics.get('followers_count', 0):,}"
                        engagement = f"{analytics.get('engagement_rate', 0):.2f}%"
                        total_posts = analytics.get('total_posts', 0)
                    else:
                        followers = "N/A"
                        engagement = "N/A"
                        total_posts = 0

                    handle = acc['profile_url'].split('/')[-1] if '/' in acc['profile_url'] else acc['profile_url']
                    markdown += f"- **Handle:** @{handle}\n"
                    markdown += f"- **Followers:** {followers}\n"
                    markdown += f"- **Engagement Rate:** {engagement}\n"
                    markdown += f"- **Total Posts:** {total_posts:,}\n"
                    markdown += f"- **Profile:** {acc['profile_url']}\n\n"

                    # Get demographics if available
                    demographics = self.db.get_demographics_data(int(acc['id']))
                    if demographics:
                        markdown += "**Audience Demographics:**\n\n"

                        if demographics.get('gender'):
                            markdown += "Gender Distribution:\n"
                            for gender, pct in demographics['gender'].items():
                                markdown += f"- {gender.title()}: {pct}%\n"
                            markdown += "\n"

                        if demographics.get('age_brackets'):
                            markdown += "Age Distribution:\n"
                            for age, pct in list(demographics['age_brackets'].items())[:5]:
                                markdown += f"- {age}: {pct}%\n"
                            markdown += "\n"

                        if demographics.get('geography'):
                            markdown += "Top Countries:\n"
                            for geo in demographics['geography'][:3]:
                                markdown += f"- {geo.get('country', 'Unknown')}: {geo.get('percentage', 0)}%\n"
                            markdown += "\n"

                    # Get posts for this platform
                    posts_df = self.db.get_posts_for_account(int(acc['id']), limit=50)
                    if not posts_df.empty:
                        # Calculate aggregate metrics
                        total_likes = posts_df['likes_count'].sum()
                        total_comments = posts_df['comments_count'].sum()
                        total_views = posts_df['views_count'].sum() if 'views_count' in posts_df else 0

                        markdown += f"**Recent Posts Analyzed:** {len(posts_df)}\n\n"
                        markdown += "Aggregate Engagement:\n"
                        markdown += f"- Total Likes: {int(total_likes):,}\n"
                        markdown += f"- Total Comments: {int(total_comments):,}\n"
                        if total_views > 0:
                            markdown += f"- Total Views: {int(total_views):,}\n"
                        markdown += "\n"

                        # Show top 3 posts
                        posts_df['engagement'] = posts_df['likes_count'] + posts_df['comments_count'] * 3
                        top_posts = posts_df.nlargest(3, 'engagement')

                        markdown += "Top 3 Posts by Engagement:\n\n"
                        for post_idx, (_, post) in enumerate(top_posts.iterrows(), 1):
                            post_date = str(post.get('post_date', ''))[:10]
                            caption = post.get('caption', '')
                            caption_preview = caption[:80] + "..." if len(caption) > 80 else caption

                            markdown += f"{post_idx}. **{post_date}** - {caption_preview}\n"
                            markdown += f"   - {int(post.get('likes_count', 0)):,} likes, {int(post.get('comments_count', 0)):,} comments"
                            if post.get('views_count', 0) > 0:
                                markdown += f", {int(post['views_count']):,} views"
                            markdown += "\n"
                            if post.get('post_url'):
                                markdown += f"   - [View Post]({post['post_url']})\n"
                            markdown += "\n"

                    markdown += "---\n\n"

            markdown += f"**Analysis Cost:** ${row['analysis_cost']:.2f}\n"
            markdown += f"**Generated:** {row['generated_at']}\n\n"
            markdown += "---\n\n"

        # Footer
        markdown += f"""
---

## Report Metadata

- **Brief Name:** {brief['name']}
- **Total Creators Analyzed:** {len(reports_df)}
- **Average Brand Fit Score:** {avg_score:.1f}/5.0
- **Total Analysis Cost:** ${total_cost:.2f}
- **Report Generated:** {datetime.now().strftime("%B %d, %Y at %I:%M %p")}

*Generated with BrandSafe Creator Analysis Platform*
"""

        return markdown

    def generate_brief_report_excel(self, brief_id: int) -> bytes:
        """
        Generate Excel report for all creators in a brief

        Args:
            brief_id: Brief ID

        Returns:
            Excel file as bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        # Get brief info
        brief = self.db.get_brief(brief_id)
        if not brief:
            return None

        # Get all reports for this brief
        reports_df = self.db.get_reports_for_brief(brief_id)
        if reports_df.empty:
            return None

        # Sort by score descending
        reports_df = reports_df.sort_values('overall_score', ascending=False)

        # Create workbook
        wb = Workbook()

        # Sheet 1: Executive Summary
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        # Add header
        ws_summary['A1'] = f"{brief['name']} - Analysis Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"

        # Add metrics
        avg_score = reports_df['overall_score'].mean()
        total_cost = reports_df['analysis_cost'].sum()

        ws_summary['A4'] = "Key Metrics"
        ws_summary['A4'].font = Font(bold=True, size=12)
        ws_summary['A5'] = "Total Creators Analyzed:"
        ws_summary['B5'] = len(reports_df)
        ws_summary['A6'] = "Average Brand Fit Score:"
        ws_summary['B6'] = f"{avg_score:.1f}/5.0"
        ws_summary['A7'] = "Total Analysis Cost:"
        ws_summary['B7'] = f"${total_cost:.2f}"

        # Top 3 recommendations
        ws_summary['A9'] = "Top 3 Recommended Creators"
        ws_summary['A9'].font = Font(bold=True, size=12)
        top_creators = reports_df.head(3)
        row_idx = 10
        for idx, (_, row) in enumerate(top_creators.iterrows(), 1):
            ws_summary[f'A{row_idx}'] = f"{idx}. {row['creator_name']}"
            ws_summary[f'B{row_idx}'] = f"{row['overall_score']:.1f}/5.0"
            ws_summary[f'C{row_idx}'] = row['primary_platform']
            row_idx += 1

        # Sheet 2: Creator Comparison
        ws_comparison = wb.create_sheet("Creator Comparison")

        # Headers
        headers = ['Rank', 'Creator', 'Platform', 'Brand Fit Score', 'Total Followers', 'Analysis Cost']
        ws_comparison.append(headers)

        # Style headers
        for cell in ws_comparison[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Add data
        for idx, (_, row) in enumerate(reports_df.iterrows(), 1):
            # Get follower count
            accounts_df = self.db.get_social_accounts(int(row['creator_id']))
            total_followers = 0
            if not accounts_df.empty:
                for _, acc in accounts_df.iterrows():
                    analytics = self.db.get_latest_analytics(int(acc['id']))
                    if analytics:
                        total_followers += analytics.get('followers_count', 0)

            ws_comparison.append([
                idx,
                row['creator_name'],
                row['primary_platform'],
                f"{row['overall_score']:.1f}/5.0",
                f"{total_followers:,}" if total_followers > 0 else "N/A",
                f"${row['analysis_cost']:.2f}"
            ])

        # Sheet 3: Detailed Reports
        ws_details = wb.create_sheet("Detailed Reports")

        # Headers
        detail_headers = ['Rank', 'Creator', 'Score', 'Summary', 'Strengths', 'Concerns', 'Recommendations']
        ws_details.append(detail_headers)

        # Style headers
        for cell in ws_details[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Add data
        for idx, (_, row) in enumerate(reports_df.iterrows(), 1):
            try:
                strengths = json.loads(row['strengths']) if isinstance(row['strengths'], str) else row['strengths']
                concerns = json.loads(row['concerns']) if isinstance(row['concerns'], str) else row['concerns']
                recommendations = json.loads(row['recommendations']) if isinstance(row['recommendations'], str) else row['recommendations']
            except:
                strengths = []
                concerns = []
                recommendations = []

            ws_details.append([
                idx,
                row['creator_name'],
                f"{row['overall_score']:.1f}/5.0",
                row.get('summary', 'N/A'),
                '\n'.join(f"• {s}" for s in strengths) if strengths else 'N/A',
                '\n'.join(f"• {c}" for c in concerns) if concerns else 'N/A',
                '\n'.join(f"• {r}" for r in recommendations) if recommendations else 'N/A'
            ])

        # Sheet 4: Post Analysis
        ws_posts = wb.create_sheet("Post Analysis")

        # Headers
        post_headers = ['Creator', 'Platform', 'Post Date', 'Post Type', 'Caption',
                        'Likes', 'Comments', 'Views', 'Brand Safety', 'Post URL']
        ws_posts.append(post_headers)

        # Style headers
        for cell in ws_posts[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Add post data for each creator
        for _, row in reports_df.iterrows():
            creator_id = int(row['creator_id'])
            creator_name = row['creator_name']
            accounts_df = self.db.get_social_accounts(creator_id)

            for _, acc in accounts_df.iterrows():
                posts_df = self.db.get_posts_for_account(int(acc['id']), limit=50)
                if not posts_df.empty:
                    # Sort by engagement
                    posts_df['engagement'] = posts_df['likes_count'] + posts_df['comments_count'] * 3
                    posts_df = posts_df.nlargest(10, 'engagement')  # Top 10 posts per platform

                    for _, post in posts_df.iterrows():
                        post_date = str(post.get('post_date', ''))[:10]
                        caption = post.get('caption', '')
                        caption_preview = caption[:100] + "..." if len(caption) > 100 else caption

                        ws_posts.append([
                            creator_name,
                            acc['platform'].title(),
                            post_date,
                            post.get('post_type', 'post'),
                            caption_preview,
                            int(post.get('likes_count', 0)),
                            int(post.get('comments_count', 0)),
                            int(post.get('views_count', 0)) if post.get('views_count') else 0,
                            f"{post.get('brand_safety_score', 0):.1f}" if post.get('brand_safety_score') else 'N/A',
                            post.get('post_url', '')
                        ])

        # Sheet 5: Demographics
        ws_demographics = wb.create_sheet("Demographics")

        # Headers
        demo_headers = ['Creator', 'Platform', 'Followers', 'Female %', 'Male %',
                        'Age 18-24 %', 'Age 25-34 %', 'Top Country', 'Top Country %', 'Data Source']
        ws_demographics.append(demo_headers)

        # Style headers
        for cell in ws_demographics[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Add demographics data
        for _, row in reports_df.iterrows():
            creator_id = int(row['creator_id'])
            creator_name = row['creator_name']
            accounts_df = self.db.get_social_accounts(creator_id)

            for _, acc in accounts_df.iterrows():
                analytics = self.db.get_latest_analytics(int(acc['id']))
                demographics = self.db.get_demographics_data(int(acc['id']))

                followers = analytics.get('followers_count', 0) if analytics else 0

                if demographics:
                    gender = demographics.get('gender', {})
                    age_brackets = demographics.get('age_brackets', {})
                    geography = demographics.get('geography', [])

                    female_pct = gender.get('female', 0)
                    male_pct = gender.get('male', 0)
                    age_18_24 = age_brackets.get('18-24', 0)
                    age_25_34 = age_brackets.get('25-34', 0)
                    top_country = geography[0].get('country', 'N/A') if geography else 'N/A'
                    top_country_pct = geography[0].get('percentage', 0) if geography else 0
                    data_source = demographics.get('data_source', 'N/A')

                    ws_demographics.append([
                        creator_name,
                        acc['platform'].title(),
                        f"{followers:,}",
                        f"{female_pct}%",
                        f"{male_pct}%",
                        f"{age_18_24}%",
                        f"{age_25_34}%",
                        top_country,
                        f"{top_country_pct}%",
                        data_source
                    ])
                else:
                    # Add row with N/A if no demographics
                    ws_demographics.append([
                        creator_name,
                        acc['platform'].title(),
                        f"{followers:,}",
                        'N/A', 'N/A', 'N/A', 'N/A', 'N/A', 'N/A', 'No data'
                    ])

        # Adjust column widths
        for ws in [ws_summary, ws_comparison, ws_details, ws_posts, ws_demographics]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_brief_report_pdf(self, brief_id: int) -> bytes:
        """
        Generate PDF report for all creators in a brief

        Args:
            brief_id: Brief ID

        Returns:
            PDF file as bytes
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")

        # Get brief info
        brief = self.db.get_brief(brief_id)
        if not brief:
            return None

        # Get all reports for this brief
        reports_df = self.db.get_reports_for_brief(brief_id)
        if reports_df.empty:
            return None

        # Sort by score descending
        reports_df = reports_df.sort_values('overall_score', ascending=False)

        # Calculate metrics
        avg_score = reports_df['overall_score'].mean()
        total_cost = reports_df['analysis_cost'].sum()

        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                                rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=18)

        # Container for the 'Flowable' objects
        elements = []

        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            spaceBefore=12
        )

        # Title
        elements.append(Paragraph(f"{brief['name']}<br/>Complete Analysis Report", title_style))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
        elements.append(Spacer(1, 24))

        # Executive Summary
        elements.append(Paragraph("Executive Summary", heading_style))
        elements.append(Paragraph(f"<b>Brief Description:</b> {brief.get('description', 'N/A')}", styles['Normal']))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"<b>Brand Context:</b> {brief.get('brand_context', 'N/A')}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Key Metrics
        metrics_data = [
            ['Metric', 'Value'],
            ['Total Creators Analyzed', str(len(reports_df))],
            ['Average Brand Fit Score', f"{avg_score:.1f}/5.0"],
            ['Total Analysis Cost', f"${total_cost:.2f}"]
        ]
        metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(metrics_table)
        elements.append(Spacer(1, 20))

        # Top 3 Recommendations
        elements.append(Paragraph("Top 3 Recommended Creators", heading_style))
        top_creators = reports_df.head(3)
        for idx, (_, row) in enumerate(top_creators.iterrows(), 1):
            score_indicator = "🟢" if row['overall_score'] >= 4.0 else "🟡" if row['overall_score'] >= 3.0 else "🔴"
            elements.append(Paragraph(f"{idx}. <b>{row['creator_name']}</b> - {row['overall_score']:.1f}/5.0 ({row['primary_platform']})", styles['Normal']))
            elements.append(Spacer(1, 6))

        elements.append(PageBreak())

        # Creator Comparison Table
        elements.append(Paragraph("Creator Comparison", heading_style))
        comparison_data = [['Rank', 'Creator', 'Platform', 'Brand Fit', 'Followers', 'Cost']]

        for idx, (_, row) in enumerate(reports_df.iterrows(), 1):
            # Get follower count
            accounts_df = self.db.get_social_accounts(int(row['creator_id']))
            total_followers = 0
            if not accounts_df.empty:
                for _, acc in accounts_df.iterrows():
                    analytics = self.db.get_latest_analytics(int(acc['id']))
                    if analytics:
                        total_followers += analytics.get('followers_count', 0)

            comparison_data.append([
                str(idx),
                row['creator_name'],
                row['primary_platform'],
                f"{row['overall_score']:.1f}/5.0",
                f"{total_followers:,}" if total_followers > 0 else "N/A",
                f"${row['analysis_cost']:.2f}"
            ])

        comparison_table = Table(comparison_data, colWidths=[0.5*inch, 1.5*inch, 1*inch, 1*inch, 1.2*inch, 0.8*inch])
        comparison_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9)
        ]))
        elements.append(comparison_table)

        elements.append(PageBreak())

        # Individual Creator Reports (enhanced for PDF)
        elements.append(Paragraph("Individual Creator Detailed Reports", heading_style))

        for idx, (_, row) in enumerate(reports_df.iterrows(), 1):
            elements.append(Paragraph(f"<b>{idx}. {row['creator_name']}</b> - {row['overall_score']:.1f}/5.0", styles['Heading3']))
            elements.append(Spacer(1, 6))
            elements.append(Paragraph(f"<b>Summary:</b> {row.get('summary', 'No summary available')}", styles['Normal']))
            elements.append(Spacer(1, 8))

            # Parse lists
            try:
                strengths = json.loads(row['strengths']) if isinstance(row['strengths'], str) else row['strengths']
                concerns = json.loads(row['concerns']) if isinstance(row['concerns'], str) else row['concerns']

                if strengths:
                    elements.append(Paragraph("<b>Key Strengths:</b>", styles['Normal']))
                    for strength in strengths[:5]:  # Top 5 strengths
                        elements.append(Paragraph(f"• {strength}", styles['Normal']))
                    elements.append(Spacer(1, 6))

                if concerns:
                    elements.append(Paragraph("<b>Potential Concerns:</b>", styles['Normal']))
                    for concern in concerns[:3]:  # Top 3 concerns
                        elements.append(Paragraph(f"• {concern}", styles['Normal']))
                    elements.append(Spacer(1, 6))
            except:
                pass

            # Platform details and posts
            creator_id = int(row['creator_id'])
            accounts_df = self.db.get_social_accounts(creator_id)

            if not accounts_df.empty:
                elements.append(Paragraph("<b>Platform Analytics:</b>", styles['Normal']))
                elements.append(Spacer(1, 4))

                for _, acc in accounts_df.iterrows():
                    platform = acc['platform'].title()
                    analytics = self.db.get_latest_analytics(int(acc['id']))

                    if analytics:
                        followers = f"{analytics.get('followers_count', 0):,}"
                        engagement = f"{analytics.get('engagement_rate', 0):.2f}%"

                        elements.append(Paragraph(f"<b>{platform}:</b> {followers} followers | {engagement} engagement", styles['Normal']))

                        # Get top 3 posts
                        posts_df = self.db.get_posts_for_account(int(acc['id']), limit=50)
                        if not posts_df.empty:
                            posts_df['engagement_calc'] = posts_df['likes_count'] + posts_df['comments_count'] * 3
                            top_posts = posts_df.nlargest(3, 'engagement_calc')

                            total_likes = int(posts_df['likes_count'].sum())
                            total_comments = int(posts_df['comments_count'].sum())

                            elements.append(Paragraph(f"  Recent Posts: {len(posts_df)} analyzed | {total_likes:,} total likes | {total_comments:,} total comments", styles['Normal']))

                            # Add top post preview
                            if not top_posts.empty:
                                top_post = top_posts.iloc[0]
                                caption = top_post.get('caption', '')
                                caption_preview = caption[:60] + "..." if len(caption) > 60 else caption
                                likes = int(top_post.get('likes_count', 0))
                                comments = int(top_post.get('comments_count', 0))
                                elements.append(Paragraph(f"  Top Post: {likes:,} likes, {comments:,} comments - {caption_preview}", styles['Normal']))

                        elements.append(Spacer(1, 4))

                elements.append(Spacer(1, 6))

            # Add page break between creators
            if idx < len(reports_df):
                elements.append(Spacer(1, 12))
                elements.append(Paragraph("─" * 80, styles['Normal']))
                elements.append(Spacer(1, 12))

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
